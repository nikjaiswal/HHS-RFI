"""Score AI codes against human reviewers (v2 sample, n=100).

Compares the LLM key (validation/v2_llm_key_sample.xlsx, sheet "LLM Key") against
the two human reviewer code tables (validation/v2_reviewer_a.csv, v2_reviewer_b.csv).
Reports % agreement, Cohen's kappa with 95% bootstrap CIs, PABAK, and Krippendorff's α
per coded variable for AI vs Reviewer A, AI vs Reviewer B, and Reviewer A vs B.
"""
from __future__ import annotations

import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
LLM_KEY_XLSX = ROOT / "validation" / "v2_llm_key_sample.xlsx"
REVIEWER_A_CSV = ROOT / "validation" / "v2_reviewer_a.csv"
REVIEWER_B_CSV = ROOT / "validation" / "v2_reviewer_b.csv"
OUT_CSV = ROOT / "output" / "ai_vs_human_v2.csv"

TOPIC_LABELS = {
    "Regulatory framework": "top_regulation",
    "Evaluation & monitoring": "top_evaluation",
    "Reimbursement & payment": "top_reimbursement",
    "Transparency & explainability": "top_transparency",
    "Clinical workflow integration": "top_workflow",
    "Trust": "top_trust",
    "Patient safety": "top_safety",
    "FDA scope & device classification": "top_fda_scope",
    "Interoperability & data standards": "top_interoperability",
    "Equity & bias": "top_equity",
    "Liability & accountability": "top_liability",
    "Admin burden reduction": "top_admin_burden",
    "Standards & accreditation": "top_standards",
    "Workforce impact": "top_workforce",
    "Privacy & HIPAA": "top_privacy",
}
BARRIER_LABELS = {
    "Regulatory uncertainty": "bar_reg_uncertainty",
    "Liability risk": "bar_liability_risk",
    "Payment misalignment": "bar_payment_misalign",
    "Data fragmentation": "bar_data_fragmentation",
    "Clinician trust deficit": "bar_clinician_trust",
    "Bias & equity risk": "bar_bias_equity",
    "Privacy constraints": "bar_privacy_constraints",
    "Cost & resource constraints": "bar_cost_resources",
}
POSITION_LABELS = {
    "Human oversight": "pos_oversight",
    "Regulatory approach": "pos_regulation",
    "Liability": "pos_liability",
    "Reimbursement": "pos_reimbursement",
    "Interoperability": "pos_interoperability",
    "Evaluation": "pos_evaluation",
}
BASE_LABELS = {
    "Commenter Type": "commenter_type",
    "Clinical Perspective": "clinical_perspective",
    "Patient Perspective": "patient_perspective",
    "Has CFR Citation": "has_cfr_citation",
    "Evidence Type": "evidence_type",
}

VAR_LABELS: dict[str, str] = {}
for d in (BASE_LABELS, TOPIC_LABELS, BARRIER_LABELS, POSITION_LABELS):
    VAR_LABELS.update(d)


def normalize(raw) -> str:
    """Reduce a coded cell to its canonical short token (e.g. '1', 'IND', 'H2', 'none')."""
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "1" if raw else "0"
    if isinstance(raw, (int, float)):
        if float(raw).is_integer():
            return str(int(raw))
        return str(raw).strip()
    s = str(raw).strip()
    if not s:
        return ""
    # Strip dropdown suffix " — ..." (em-dash) or " � ..." (mojibake replacement char)
    for sep in (" — ", " — ", " � ", " - "):
        if sep in s:
            return s.split(sep, 1)[0].strip()
    # Fall back to first whitespace token (canonical codes are single tokens)
    return s.split()[0]


def load_llm_key() -> dict[str, dict[str, str]]:
    wb = load_workbook(LLM_KEY_XLSX, data_only=True, read_only=True)
    ws = wb["LLM Key"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [h for h in rows[0]]
    out: dict[str, dict[str, str]] = {}
    cid_idx = headers.index("Comment ID")
    for row in rows[1:]:
        cid = row[cid_idx]
        if not cid:
            continue
        record: dict[str, str] = {}
        for label, var in VAR_LABELS.items():
            if label not in headers:
                continue
            record[var] = normalize(row[headers.index(label)])
        out[cid] = record
    return out


def _open_csv(path: Path):
    """Open a CSV trying utf-8-sig first, then cp1252 (some reviewer exports are Windows-1252)."""
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as probe:
                probe.read()
            return open(path, "r", encoding=enc, newline="")
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Could not decode {path} with any known encoding")


def load_reviewer_csv(path: Path) -> dict[str, dict[str, str]]:
    """Load a reviewer code table.

    Canonical format (v2_reviewer_a.csv, v2_reviewer_b.csv): columns are
    `comment_id` plus one column per coded variable, values already canonical.
    """
    out: dict[str, dict[str, str]] = {}
    with _open_csv(path) as f:
        reader = csv.DictReader(f)
        for row in reader:
            cid = (row.get("comment_id") or row.get("Comment ID") or "").strip()
            if not cid:
                continue
            record: dict[str, str] = {}
            for var in VAR_LABELS.values():
                if var in row:
                    record[var] = normalize(row[var])
            out[cid] = record
    return out


def cohens_kappa(values_a: list[str], values_b: list[str]) -> tuple[float, float, float]:
    """Return (pct_agreement, kappa, pabak).

    PABAK (prevalence-adjusted bias-adjusted kappa) = 2*Po - 1, valid for binary variables.
    For multi-class variables we still report it; interpret with caution.
    """
    n = len(values_a)
    if n == 0:
        return 0.0, 0.0, 0.0
    agree = sum(1 for a, b in zip(values_a, values_b) if a == b)
    po = agree / n
    counts_a: dict[str, int] = defaultdict(int)
    counts_b: dict[str, int] = defaultdict(int)
    for a, b in zip(values_a, values_b):
        counts_a[a] += 1
        counts_b[b] += 1
    cats = set(counts_a) | set(counts_b)
    pe = sum((counts_a[c] / n) * (counts_b[c] / n) for c in cats)
    if pe >= 1.0:
        kappa = 1.0 if po == 1.0 else 0.0
    else:
        kappa = (po - pe) / (1 - pe)
    pabak = 2 * po - 1
    return po * 100, kappa, pabak


def bootstrap_kappa_ci(
    values_a: list[str],
    values_b: list[str],
    n_iter: int = 2000,
    seed: int = 42,
    alpha: float = 0.05,
) -> tuple[float, float]:
    """Percentile bootstrap 95% CI on Cohen's κ. Returns (lo, hi)."""
    import random
    n = len(values_a)
    if n == 0:
        return 0.0, 0.0
    rng = random.Random(seed)
    indices = list(range(n))
    estimates = []
    for _ in range(n_iter):
        sample = [rng.choice(indices) for _ in range(n)]
        a_sample = [values_a[i] for i in sample]
        b_sample = [values_b[i] for i in sample]
        _, k, _ = cohens_kappa(a_sample, b_sample)
        estimates.append(k)
    estimates.sort()
    lo = estimates[int((alpha / 2) * n_iter)]
    hi = estimates[int((1 - alpha / 2) * n_iter)]
    return lo, hi


def krippendorffs_alpha(values_a: list[str], values_b: list[str]) -> float:
    """Krippendorff's α for nominal data, 2 raters. Returns α in [-1, 1].

    Implementation of the standard nominal-α formula; for 2 raters this reduces to
    α = 1 - (D_o / D_e), where disagreements are computed from the coincidence matrix.
    """
    n = len(values_a)
    if n == 0:
        return 0.0
    # Categories
    cats = sorted(set(values_a) | set(values_b))
    if len(cats) < 2:
        return 1.0  # Everyone agrees on one category, perfect by definition.
    cat_idx = {c: i for i, c in enumerate(cats)}
    K = len(cats)

    # Coincidence matrix C[k,l] for nominal data (2 raters): unordered pair counts.
    C = [[0.0 for _ in range(K)] for _ in range(K)]
    for a, b in zip(values_a, values_b):
        i, j = cat_idx[a], cat_idx[b]
        # 2 raters: each unit contributes 2 coincidences (a-b and b-a),
        # divided by (number of raters - 1) = 1, summed over rater pairs.
        # Simpler: for each pair, increment C[i,j] and C[j,i] by 1.
        C[i][j] += 1.0
        C[j][i] += 1.0

    n_total = sum(sum(row) for row in C)  # = 2n
    n_k = [sum(row) for row in C]  # marginal totals

    if n_total == 0:
        return 0.0

    # Observed disagreements (off-diagonal sum, weighted equally for nominal)
    D_o = 0.0
    for i in range(K):
        for j in range(K):
            if i != j:
                D_o += C[i][j]
    D_o /= n_total

    # Expected disagreements
    D_e = 0.0
    for i in range(K):
        for j in range(K):
            if i != j:
                D_e += n_k[i] * n_k[j]
    if n_total > 1:
        D_e /= n_total * (n_total - 1)
    else:
        return 0.0

    if D_e == 0:
        return 1.0
    return 1 - D_o / D_e


MULTICLASS_VARS = {"commenter_type", "evidence_type", "pos_oversight", "pos_regulation",
                   "pos_liability", "pos_reimbursement", "pos_interoperability", "pos_evaluation"}


def prevalence_yes(values: list[str]) -> float:
    """For binary vars, fraction coded as '1'. Returns 0 for non-binary; check var separately."""
    if not values:
        return 0.0
    return sum(1 for v in values if v == "1") / len(values)


def aligned_values(
    src_a: dict[str, dict[str, str]],
    src_b: dict[str, dict[str, str]],
    var: str,
    ids: list[str],
) -> tuple[list[str], list[str]]:
    a_vals, b_vals = [], []
    for cid in ids:
        a_vals.append(src_a.get(cid, {}).get(var, ""))
        b_vals.append(src_b.get(cid, {}).get(var, ""))
    return a_vals, b_vals


def main() -> None:
    ai = load_llm_key()
    ra = load_reviewer_csv(REVIEWER_A_CSV)
    rb = load_reviewer_csv(REVIEWER_B_CSV)

    common_ids = sorted(set(ai) & set(ra) & set(rb))
    print(
        f"AI key: {len(ai)} | Reviewer A: {len(ra)} | Reviewer B: {len(rb)} "
        f"| common: {len(common_ids)}",
        file=sys.stderr,
    )

    print("Computing per-variable κ + bootstrap CIs (2000 iterations) ...", file=sys.stderr)
    rows = []
    for var in sorted(VAR_LABELS.values()):
        a_ai, a_ra = aligned_values(ai, ra, var, common_ids)
        _, a_rb = aligned_values(ai, rb, var, common_ids)
        ra_vals, rb_vals = aligned_values(ra, rb, var, common_ids)

        ai_a_pct, ai_a_k, ai_a_pabak = cohens_kappa(a_ai, a_ra)
        ai_b_pct, ai_b_k, ai_b_pabak = cohens_kappa(a_ai, a_rb)
        ab_pct, ab_k, ab_pabak = cohens_kappa(ra_vals, rb_vals)

        ai_a_lo, ai_a_hi = bootstrap_kappa_ci(a_ai, a_ra)
        ai_b_lo, ai_b_hi = bootstrap_kappa_ci(a_ai, a_rb)
        ab_lo, ab_hi = bootstrap_kappa_ci(ra_vals, rb_vals)

        # Krippendorff's alpha
        ai_a_alpha = krippendorffs_alpha(a_ai, a_ra)
        ai_b_alpha = krippendorffs_alpha(a_ai, a_rb)
        ab_alpha = krippendorffs_alpha(ra_vals, rb_vals)

        is_binary = var not in MULTICLASS_VARS
        prev_ai = prevalence_yes(a_ai) if is_binary else None
        prev_a = prevalence_yes(a_ra) if is_binary else None
        prev_b = prevalence_yes(a_rb) if is_binary else None

        rows.append(
            {
                "variable": var,
                "type": "multiclass" if not is_binary else "binary",
                "n": len(common_ids),
                "prev_ai": round(prev_ai, 3) if prev_ai is not None else "",
                "prev_a": round(prev_a, 3) if prev_a is not None else "",
                "prev_b": round(prev_b, 3) if prev_b is not None else "",
                "ai_vs_a_pct": round(ai_a_pct, 2),
                "ai_vs_a_kappa": round(ai_a_k, 4),
                "ai_vs_a_kappa_lo": round(ai_a_lo, 4),
                "ai_vs_a_kappa_hi": round(ai_a_hi, 4),
                "ai_vs_a_pabak": round(ai_a_pabak, 4),
                "ai_vs_a_alpha": round(ai_a_alpha, 4),
                "ai_vs_b_pct": round(ai_b_pct, 2),
                "ai_vs_b_kappa": round(ai_b_k, 4),
                "ai_vs_b_kappa_lo": round(ai_b_lo, 4),
                "ai_vs_b_kappa_hi": round(ai_b_hi, 4),
                "ai_vs_b_pabak": round(ai_b_pabak, 4),
                "ai_vs_b_alpha": round(ai_b_alpha, 4),
                "a_vs_b_pct": round(ab_pct, 2),
                "a_vs_b_kappa": round(ab_k, 4),
                "a_vs_b_kappa_lo": round(ab_lo, 4),
                "a_vs_b_kappa_hi": round(ab_hi, 4),
                "a_vs_b_pabak": round(ab_pabak, 4),
                "a_vs_b_alpha": round(ab_alpha, 4),
            }
        )

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nWrote {OUT_CSV.relative_to(ROOT)} ({len(rows)} variables)\n")

    # Console summary
    width = max(len(r["variable"]) for r in rows)
    header = (
        f"{'variable':<{width}}  n   AI–A%  κ_AI-A   AI–B%  κ_AI-B   A–B%  κ_A-B"
    )
    print(header)
    print("-" * len(header))
    for r in rows:
        print(
            f"{r['variable']:<{width}}  {r['n']:>3}  "
            f"{r['ai_vs_a_pct']:>5.1f}  {r['ai_vs_a_kappa']:>+6.3f}  "
            f"{r['ai_vs_b_pct']:>5.1f}  {r['ai_vs_b_kappa']:>+6.3f}  "
            f"{r['a_vs_b_pct']:>5.1f}  {r['a_vs_b_kappa']:>+6.3f}"
        )

    # Macro means
    def avg(field: str) -> float:
        return sum(r[field] for r in rows) / len(rows)

    print()
    print(
        f"MEAN          ─    "
        f"{avg('ai_vs_a_pct'):>5.1f}  {avg('ai_vs_a_kappa'):>+6.3f}  "
        f"{avg('ai_vs_b_pct'):>5.1f}  {avg('ai_vs_b_kappa'):>+6.3f}  "
        f"{avg('a_vs_b_pct'):>5.1f}  {avg('a_vs_b_kappa'):>+6.3f}"
    )

    # Summary statistics for manuscript reporting
    print(
        f"\nManuscript summary:\n"
        f"  AI vs Reviewer A:  mean κ = {avg('ai_vs_a_kappa'):.3f}, "
        f"mean PABAK = {avg('ai_vs_a_pabak'):.3f}, "
        f"mean agreement = {avg('ai_vs_a_pct'):.1f}%\n"
        f"  AI vs Reviewer B:  mean κ = {avg('ai_vs_b_kappa'):.3f}, "
        f"mean PABAK = {avg('ai_vs_b_pabak'):.3f}, "
        f"mean agreement = {avg('ai_vs_b_pct'):.1f}%\n"
        f"  Reviewer A vs B:   mean κ = {avg('a_vs_b_kappa'):.3f}, "
        f"mean PABAK = {avg('a_vs_b_pabak'):.3f}, "
        f"mean agreement = {avg('a_vs_b_pct'):.1f}%"
    )


if __name__ == "__main__":
    main()
